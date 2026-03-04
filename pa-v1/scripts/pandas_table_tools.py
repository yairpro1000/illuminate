#!/usr/bin/env python3
"""
CSV/XLSX utilities (pandas) for future PA v2 workflows.

Not wired into the app. Intended as a standalone helper for:
- Inspecting tables
- Converting CSV <-> XLSX
- Exporting/importing PA list JSONL <-> CSV/XLSX (schema-aware)
"""

from __future__ import annotations

import argparse
import json
import sys
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Literal

import pandas as pd


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _excel_argb_from_hex_color(color: Any, alpha_hex: str = "33") -> str | None:
    if not isinstance(color, str):
        return None
    h = color.strip().lstrip("#")
    if len(h) != 6:
        return None
    try:
        int(h, 16)
    except ValueError:
        return None
    # ARGB (alpha + RGB). Alpha "33" is a subtle tint (similar to the app export).
    return f"{alpha_hex}{h}".upper()


def _is_missing(value: Any) -> bool:
    # pandas uses NaN/NaT for missing values; pd.isna handles scalars well.
    try:
        return bool(pd.isna(value))
    except Exception:
        return value is None


def _read_table(path: Path, *, sheet: str | int | None = None) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix in SUPPORTED_EXCEL_SUFFIXES:
        return pd.read_excel(path, sheet_name=sheet if sheet is not None else 0, engine="openpyxl")
    raise ValueError(f"Unsupported input type: {path.name} (expected .csv, .xlsx, or .xlsm)")


def _write_table(df: pd.DataFrame, path: Path, *, sheet_name: str = "Sheet1") -> None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        df.to_csv(path, index=False)
        return
    if suffix in SUPPORTED_EXCEL_SUFFIXES:
        # Use an explicit writer so we can style rows (v2 prep).
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.book[sheet_name]

            # If there is a `color` column with hex values (#RRGGBB), tint the entire row.
            if "color" in df.columns:
                from openpyxl.styles import PatternFill

                color_col_idx = int(df.columns.get_loc("color")) + 1  # 1-based in Excel
                max_col = int(ws.max_column)
                for excel_row in range(2, len(df) + 2):  # row 1 is header
                    cell_color = ws.cell(row=excel_row, column=color_col_idx).value
                    argb = _excel_argb_from_hex_color(cell_color)
                    if not argb:
                        continue
                    fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
                    for c in range(1, max_col + 1):
                        ws.cell(row=excel_row, column=c).fill = fill
        return
    raise ValueError(f"Unsupported output type: {path.name} (expected .csv, .xlsx, or .xlsm)")


def cmd_inspect(args: argparse.Namespace) -> int:
    df = _read_table(Path(args.input), sheet=args.sheet)
    print(f"Rows: {len(df)}")
    print(f"Cols: {len(df.columns)}")
    print("Columns:")
    for c in df.columns:
        print(f"- {c} ({df[c].dtype})")
    head_n = int(args.head)
    if head_n > 0:
        print()
        print(df.head(head_n).to_string(index=False))
    return 0


def cmd_convert(args: argparse.Namespace) -> int:
    inp = Path(args.input)
    out = Path(args.output)
    df = _read_table(inp, sheet=args.sheet)
    _write_table(df, out, sheet_name=args.out_sheet or "Sheet1")
    return 0


@dataclass(frozen=True)
class FieldDef:
    name: str
    type: Literal["string", "int", "float", "boolean", "date", "time", "json"]
    default: Any | None
    nullable: bool


@dataclass(frozen=True)
class ListDef:
    list_id: str
    title: str
    fields: list[FieldDef]


def _load_schema(schema_path: Path) -> dict[str, Any]:
    with schema_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _load_list_def(schema: dict[str, Any], list_id: str) -> ListDef:
    lists = schema.get("lists") or {}
    raw = lists.get(list_id)
    if not raw:
        raise ValueError(f'List "{list_id}" not found in schema.')
    title = str(raw.get("title") or list_id)
    raw_fields: dict[str, Any] = raw.get("fields") or {}

    fields: list[FieldDef] = []
    for name, fdef in raw_fields.items():
        ftype = str(fdef.get("type"))
        if ftype not in {"string", "int", "float", "boolean", "date", "time", "json"}:
            raise ValueError(f'Unsupported field type "{ftype}" for field "{name}".')
        fields.append(
            FieldDef(
                name=name,
                type=ftype,  # type: ignore[assignment]
                default=fdef.get("default", None),
                nullable=bool(fdef.get("nullable", False)),
            )
        )
    return ListDef(list_id=list_id, title=title, fields=fields)


def _iter_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            s = line.strip()
            if not s:
                continue
            try:
                obj = json.loads(s)
            except json.JSONDecodeError as e:
                raise ValueError(f"Invalid JSON on line {line_no} in {path.name}: {e}") from e
            if not isinstance(obj, dict):
                raise ValueError(f"Expected object JSON on line {line_no} in {path.name}.")
            yield obj


def _coerce_boolean(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(int(value))
    if isinstance(value, str):
        s = value.strip().lower()
        if s in {"true", "t", "yes", "y", "1"}:
            return True
        if s in {"false", "f", "no", "n", "0"}:
            return False
    raise ValueError(f"Invalid boolean value: {value!r}")


def _coerce_value(value: Any, fdef: FieldDef) -> Any:
    if _is_missing(value):
        if fdef.default is not None:
            return fdef.default
        if fdef.nullable:
            return None
        return None

    t = fdef.type
    if t == "string":
        return str(value)
    if t == "int":
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int,)):
            return int(value)
        if isinstance(value, (float,)) and value.is_integer():
            return int(value)
        if isinstance(value, str) and value.strip() != "":
            return int(float(value.strip()))
        raise ValueError(f"Invalid int value for {fdef.name}: {value!r}")
    if t == "float":
        if isinstance(value, bool):
            return float(int(value))
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str) and value.strip() != "":
            return float(value.strip())
        raise ValueError(f"Invalid float value for {fdef.name}: {value!r}")
    if t == "boolean":
        return _coerce_boolean(value)
    if t == "date":
        # Store as ISO date string (YYYY-MM-DD).
        if isinstance(value, str):
            return value.strip()
        if hasattr(value, "date"):
            return value.date().isoformat()  # pandas Timestamp, datetime
        raise ValueError(f"Invalid date value for {fdef.name}: {value!r}")
    if t == "time":
        # Store as ISO time string (HH:MM:SS).
        if isinstance(value, str):
            return value.strip()
        if hasattr(value, "time"):
            return value.time().isoformat(timespec="seconds")  # pandas Timestamp, datetime
        raise ValueError(f"Invalid time value for {fdef.name}: {value!r}")
    if t == "json":
        if isinstance(value, (dict, list)):
            return value
        if isinstance(value, str):
            s = value.strip()
            if s == "":
                return None if fdef.nullable else None
            try:
                return json.loads(s)
            except json.JSONDecodeError as e:
                raise ValueError(f"Invalid JSON value for {fdef.name}: {value!r}") from e
        # Fall back to best-effort JSON serialization
        return value
    raise ValueError(f"Unhandled field type: {t}")


def _pa_paths(data_dir: Path, list_id: str) -> tuple[Path, Path]:
    schema_path = data_dir / "meta" / "lists.schema.json"
    list_path = data_dir / "lists" / f"{list_id}.jsonl"
    return schema_path, list_path


def cmd_pa_export(args: argparse.Namespace) -> int:
    data_dir = Path(args.data_dir)
    list_id = str(args.list_id)
    schema_path, list_path = _pa_paths(data_dir, list_id)

    schema = _load_schema(schema_path)
    ldef = _load_list_def(schema, list_id)
    items = list(_iter_jsonl(list_path))

    headers = ["id", "createdAt", *[f.name for f in ldef.fields]]
    rows: list[dict[str, Any]] = []
    for it in items:
        row: dict[str, Any] = {}
        for h in headers:
            row[h] = it.get(h, None)
        rows.append(row)

    df = pd.DataFrame(rows, columns=headers)
    out = Path(args.output)
    _write_table(df, out, sheet_name=ldef.title[:31] if ldef.title else "Sheet1")
    return 0


def cmd_pa_import(args: argparse.Namespace) -> int:
    data_dir = Path(args.data_dir)
    list_id = str(args.list_id)
    schema_path, _list_path = _pa_paths(data_dir, list_id)

    schema = _load_schema(schema_path)
    ldef = _load_list_def(schema, list_id)

    df = _read_table(Path(args.input), sheet=args.sheet)
    df_cols = {str(c) for c in df.columns}

    reserved = {"id", "createdAt"}
    field_names = {f.name for f in ldef.fields}

    extra_cols = sorted([c for c in df_cols if c not in reserved and c not in field_names])
    if extra_cols and not args.allow_extra_cols:
        raise ValueError(
            "Input has columns not present in schema (use --allow-extra-cols to keep them): "
            + ", ".join(extra_cols)
        )

    out_items: list[dict[str, Any]] = []
    for _idx, row in df.iterrows():
        item: dict[str, Any] = {}

        raw_id = row.get("id", None)
        item["id"] = str(raw_id).strip() if not _is_missing(raw_id) and str(raw_id).strip() else str(uuid.uuid4())

        raw_created = row.get("createdAt", None)
        item["createdAt"] = str(raw_created).strip() if not _is_missing(raw_created) else _now_iso()

        for fdef in ldef.fields:
            if fdef.name in df.columns:
                item[fdef.name] = _coerce_value(row.get(fdef.name, None), fdef)
            else:
                item[fdef.name] = fdef.default if fdef.default is not None else (None if fdef.nullable else None)

        if args.allow_extra_cols:
            for c in extra_cols:
                item[c] = row.get(c, None)

        out_items.append(item)

    if args.dry_run:
        print(f"Validated {len(out_items)} rows for list '{list_id}'. (dry-run)")
        return 0

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        for it in out_items:
            f.write(json.dumps(it, ensure_ascii=False) + "\n")
    return 0


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="pandas_table_tools", add_help=True)
    sub = p.add_subparsers(dest="cmd", required=True)

    p_inspect = sub.add_parser("inspect", help="Print basic info + head() for a CSV/XLSX.")
    p_inspect.add_argument("input", help="Input .csv/.xlsx/.xlsm file")
    p_inspect.add_argument("--sheet", help="Excel sheet name or index", default=None)
    p_inspect.add_argument("--head", help="Number of rows to print", default=5)
    p_inspect.set_defaults(func=cmd_inspect)

    p_convert = sub.add_parser("convert", help="Convert CSV <-> XLSX using pandas.")
    p_convert.add_argument("input", help="Input .csv/.xlsx/.xlsm file")
    p_convert.add_argument("output", help="Output .csv/.xlsx/.xlsm file")
    p_convert.add_argument("--sheet", help="Excel sheet name or index (for input)", default=None)
    p_convert.add_argument("--out-sheet", help="Excel sheet name (for output)", default=None)
    p_convert.set_defaults(func=cmd_convert)

    p_pa_export = sub.add_parser("pa-export", help="Export PA list JSONL to CSV/XLSX (schema-aware).")
    p_pa_export.add_argument("--data-dir", default=str(Path(__file__).resolve().parents[1] / "data"))
    p_pa_export.add_argument("--list-id", required=True)
    p_pa_export.add_argument("--output", required=True, help="Output .csv/.xlsx/.xlsm file")
    p_pa_export.set_defaults(func=cmd_pa_export)

    p_pa_import = sub.add_parser("pa-import", help="Import CSV/XLSX to PA list JSONL (schema-aware).")
    p_pa_import.add_argument("--data-dir", default=str(Path(__file__).resolve().parents[1] / "data"))
    p_pa_import.add_argument("--list-id", required=True)
    p_pa_import.add_argument("--input", required=True, help="Input .csv/.xlsx/.xlsm file")
    p_pa_import.add_argument("--output", required=True, help="Output .jsonl file to write (does not modify app data automatically)")
    p_pa_import.add_argument("--sheet", help="Excel sheet name or index", default=None)
    p_pa_import.add_argument("--dry-run", action="store_true", help="Validate only; do not write output")
    p_pa_import.add_argument(
        "--allow-extra-cols",
        action="store_true",
        help="Keep columns not present in schema (stored as-is in JSONL)",
    )
    p_pa_import.set_defaults(func=cmd_pa_import)

    return p


def main(argv: list[str]) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return int(args.func(args))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
